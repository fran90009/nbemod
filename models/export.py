"""
N-BeMod — Excel Export Builder
Genera un fichero Excel con tabs: curves | cashflows | metadata
"""
import io
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


DARK = "0D0D0D"
ACCENT = "00C2FF"
HEADER_BG = "111827"
HEADER_FG = "FFFFFF"
ALT_ROW = "F0F4F8"


def _style_header_row(ws, row_idx: int, n_cols: int):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(bold=True, color=HEADER_FG, name="Calibri", size=10)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color=ACCENT))


def _write_df_to_sheet(ws, df: pd.DataFrame, title: str = ""):
    if title:
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12, color=DARK, name="Calibri")
        start_row = 3
    else:
        start_row = 1

    for c_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=c_idx, value=col_name)
    _style_header_row(ws, start_row, len(df.columns))

    for r_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = Font(name="Calibri", size=9)
            if r_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ALT_ROW)

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0 for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


def build_excel(
    curves_df: pd.DataFrame,
    cashflows_df: pd.DataFrame,
    config: dict,
    run_id: str,
) -> bytes:
    wb = openpyxl.Workbook()

    # ─── Sheet: curves ────────────────────────────────────────────────────────
    ws_curves = wb.active
    ws_curves.title = "curves"
    _write_df_to_sheet(ws_curves, curves_df, "CPR / SMM Curves by Segment")

    # ─── Sheet: cashflows ─────────────────────────────────────────────────────
    ws_cf = wb.create_sheet("cashflows")
    _write_df_to_sheet(ws_cf, cashflows_df, "Monthly Cashflows by Segment")

    # ─── Sheet: metadata ──────────────────────────────────────────────────────
    ws_meta = wb.create_sheet("metadata")
    ws_meta.title = "metadata"
    ws_meta.cell(row=1, column=1, value="N-BeMod Export").font = Font(bold=True, size=14, name="Calibri")
    ws_meta.cell(row=2, column=1, value=f"Generated: {datetime.utcnow().isoformat()} UTC").font = Font(size=9, color="777777")

    meta_rows = [
        ("run_id", run_id),
        ("model", "prepay_curve"),
        ("curve_method", config.get("curve_method", "simple_average")),
        ("horizon_months", config.get("horizon_months", 60)),
        ("smoothing", config.get("smoothing", False)),
        ("segments", curves_df["segment"].nunique() if not curves_df.empty else 0),
        ("total_cashflow_rows", len(cashflows_df)),
        ("total_prepayment", round(cashflows_df["prepayment"].sum(), 2) if not cashflows_df.empty else 0),
        ("export_timestamp", datetime.utcnow().isoformat()),
    ]
    _style_header_row(ws_meta, 4, 2)
    ws_meta.cell(row=4, column=1, value="Parameter")
    ws_meta.cell(row=4, column=2, value="Value")

    for i, (k, v) in enumerate(meta_rows, 5):
        ws_meta.cell(row=i, column=1, value=k).font = Font(name="Calibri", size=9, bold=True)
        ws_meta.cell(row=i, column=2, value=v).font = Font(name="Calibri", size=9)

    ws_meta.column_dimensions["A"].width = 30
    ws_meta.column_dimensions["B"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
